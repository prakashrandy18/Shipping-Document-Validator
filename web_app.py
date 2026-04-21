"""
Shipping Document Validator - Web Application
==============================================
A modern web application to compare shipping details across THREE PDF documents.
Powered by Google Gemini 1.5 Pro (Multimodal).
"""

from flask import Flask, render_template, request, jsonify
import os
import tempfile
import time
import concurrent.futures
import zipfile
import glob
import shutil
import csv
import io
from flask import Flask, render_template, request, jsonify, send_file
from dotenv import load_dotenv
from shipping_logic import extract_combined_shipping_details_llm, extract_shipping_details_llm, compare_three_documents, classify_document, GENAI_AVAILABLE, GOOGLE_API_KEY


import logging

# Configure Logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 1024 * 1024 * 1024  # Increased to 1GB for large batches
app.config['UPLOAD_FOLDER'] = tempfile.gettempdir()


@app.route('/')
def index():
    """Render the main page."""
    return render_template('index.html')











# --- Async Job Management ---
import uuid
import threading

# Job Store (In-memory for simplicity)
JOBS = {}


from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

def process_batch_job(job_id, file_paths, app_instance):
    """
    Background worker to process ZIP files and generate Excel report.
    """
    JOBS[job_id]['status'] = 'processing'
    JOBS[job_id]['progress'] = 0
    JOBS[job_id]['total'] = len(file_paths)
    
    results = []
    
    # Create temp dir for this job (already exists if passed from main, but ensure structure)
    # We used to make a temp dir here, but now we use the one where files are saved or a new one?
    # Actually, let's keep using the directory where files are as the working dir, 
    # OR create a dedicated output dir.
    
    # Let's assume file_paths are already in a dedicated job directory.
    job_dir = os.path.dirname(file_paths[0]) if file_paths else tempfile.mkdtemp()
    
    # Create dir for Renamed BLs
    renamed_bls_dir = os.path.join(job_dir, "Renamed_BLs")
    os.makedirs(renamed_bls_dir, exist_ok=True)
    
    logger.info(f"Job {job_id}: Started processing {len(file_paths)} files.")
    
    try:
        # We can process ZIPs in parallel too! 
        # Reverted workers to 5 for better throughput (7 triggered rate limits)
        with concurrent.futures.ThreadPoolExecutor(max_workers=5) as executor:
            future_to_zip = {}
            
            # 1. Submit all Zips / PDFs
            for file_path in file_paths:
                # Check extension
                filename = os.path.basename(file_path)
                if file_path.lower().endswith('.pdf'):
                     future_to_zip[executor.submit(process_combined_pdf, file_path, renamed_bls_dir)] = filename
                else:
                     future_to_zip[executor.submit(process_single_zip, file_path, renamed_bls_dir)] = filename
            
            # 2. Collect Results
            completed_count = 0
            for future in concurrent.futures.as_completed(future_to_zip):
                zip_name = future_to_zip[future]
                try:
                    res = future.result() # Returns a list of rows (usually 1 row per zip)
                    results.extend(res)
                    logger.info(f"Job {job_id}: Processed {zip_name} - Status: {res[0].get('Status')}")
                except Exception as e:
                    logger.error(f"Job {job_id}: Error processing {zip_name}: {e}")
                    results.append({'Zip_Filename': zip_name, 'Status': 'Error', 'Error_Message': str(e)})
                
                completed_count += 1
                JOBS[job_id]['progress'] = int((completed_count / len(file_paths)) * 100)
        
        # 3. Generate Excel Report using openpyxl
        wb = Workbook()
        ws = wb.active
        ws.title = "Batch Report"

        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid") # Indigo
        
        error_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid") # Red-ish light
        error_font = Font(color="991B1B") # Dark Red text

        center_align = Alignment(horizontal="center", vertical="center")
        left_align = Alignment(horizontal="left", vertical="center")
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        # Header Row
        headers = ["ZIP FILE", "STATUS", "ERRORS", "FIELD", "OBL/PKL (Doc A)", "INVOICE (Doc B)", "PACKING LIST (Doc C)"]
        ws.append(headers)
        
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border

        # Data Rows
        row_idx = 2
        for r in results:
            # We want to group by Zip. 
            # Structure: 
            # Row 1: Zip Name | Status | Errors (Merged across?)
            # Row 2-4: Details (Cartons, Weight, Volume)
            
            is_match = r.get('Status') == 'MATCH'
            
            # --- Main Zip Info Row ---
            # Zip Filename
            ws.cell(row=row_idx, column=1, value=r.get('Zip_Filename')).alignment = left_align
            
            # Status
            status_cell = ws.cell(row=row_idx, column=2, value=r.get('Status'))
            status_cell.alignment = center_align
            
            # Error Msg
            error_cell = ws.cell(row=row_idx, column=3, value=r.get('Error_Message', ''))
            error_cell.alignment = left_align
            
            # Apply Highlight if Mismatch
            if not is_match:
                for col in range(1, 8): # Approx columns
                    c = ws.cell(row=row_idx, column=col)
                    c.fill = error_fill
                    c.font = error_font
                    c.border = thin_border
            else:
                 # Standard border
                 for col in range(1, 8):
                    ws.cell(row=row_idx, column=col).border = thin_border

            row_idx += 1

            # --- Detail Rows (Optional, but useful for user to see WHAT failed) ---
            # Using data from r
            doc_a = r.get('doc_a_Name', 'Doc A')
            doc_b = r.get('doc_b_Name', 'Doc B')
            doc_c = r.get('doc_c_Name', 'Doc C')
            
            # Helper to get value
            def g(k): return str(r.get(k) or '--')

            # Fields to show: Cartons, Weight, Volume
            fields = [
                ("Cartons", 'doc_a_Cartons', 'doc_b_Cartons', 'doc_c_Cartons'),
                ("Gross Weight", 'doc_a_Weight', 'doc_b_Weight', 'doc_c_Weight'),
                ("Volume (CBM)", 'doc_a_Volume', 'doc_b_Volume', 'doc_c_Volume'),
            ]

            for label, ka, kb, kc in fields:
                ws.cell(row=row_idx, column=4, value=label).alignment = left_align
                ws.cell(row=row_idx, column=5, value=g(ka)).alignment = center_align
                ws.cell(row=row_idx, column=6, value=g(kb)).alignment = center_align
                ws.cell(row=row_idx, column=7, value=g(kc)).alignment = center_align
                
                # Check consistency for THIS field to maybe highlight specifically? 
                # For now, just inherit the "Red" feel if the whole zip is bad, OR keep it plain. 
                # User asked: "mismatch field to hghlight on red". 
                # So if this specific field has mismatch, we should highlight it.
                # However, our 'res' object doesn't explicitly flag WHICH field failed in a distinct key for easy access here (it's in 'comparisons' inside the parallel worker which is flattened to 'r').
                # We have 'Error_Message' containing text like "Gross Weight (KGS) error;". We can fuzzy check.
                
                field_error = label.split(' ')[0] in (r.get('Error_Message') or '') # Simple heuristic
                if not is_match and field_error:
                     for col in range(4, 8):
                        c = ws.cell(row=row_idx, column=col)
                        c.fill = error_fill
                        c.font = error_font
                
                # Borders
                for col in range(1, 8):
                    ws.cell(row=row_idx, column=col).border = thin_border
                
                row_idx += 1
            
            # Empty spacer row
            row_idx += 1

        # Adjust Columns
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 40
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 20
        ws.column_dimensions['F'].width = 20
        ws.column_dimensions['G'].width = 20

        # Save to Buffer
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        # Store Result
        # Zip the Renamed BLs folder
        zip_buffer = io.BytesIO()
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            for root, dirs, files in os.walk(renamed_bls_dir):
                for file in files:
                    file_path = os.path.join(root, file)
                    arcname = os.path.relpath(file_path, renamed_bls_dir)
                    zip_file.write(file_path, arcname)
        
        JOBS[job_id]['bl_zip_data'] = zip_buffer.getvalue()
        JOBS[job_id]['xlsx_data'] = output.getvalue()
        JOBS[job_id]['results'] = results  # Store detailed results for UI
        JOBS[job_id]['status'] = 'completed'
        
    except Exception as e:
        JOBS[job_id]['status'] = 'failed'
        JOBS[job_id]['error'] = str(e)
        logger.error(f"Job {job_id} failed completely: {e}")
        print(f"Job {job_id} failed: {e}")
    finally:
        # Clean up job directory?
        # Maybe keep it for a bit or rely on OS temp cleaning
        # For now, let's remove it to save space, but AFTER serving files?
        # BUT: we serve files from memory (BytesIO) in this code, so removing dir is fine.
        shutil.rmtree(job_dir, ignore_errors=True)

def process_single_zip(zip_path, renamed_bls_dir=None):
    """
    Helper to process ONE zip file. Returns a list of result rows (usually 1).
    """
    start_time = time.time()
    row = {'Zip_Filename': os.path.basename(zip_path), 'Status': '', 'Error_Message': ''}
    temp_extract_dir = tempfile.mkdtemp()
    
    try:
        with zipfile.ZipFile(zip_path, 'r') as zip_ref:
            zip_ref.extractall(temp_extract_dir)
        
        pdfs = glob.glob(os.path.join(temp_extract_dir, "**", "*.[pP][dD][fF]"), recursive=True)
        pdfs = [p for p in pdfs if not os.path.basename(p).startswith('.')]
        
        if len(pdfs) < 2:
            row['Status'] = 'Skipped'
            row['Error_Message'] = f"Found {len(pdfs)} PDFs (Need 2+)"
            return [row]
        
        # Classification
        assigned_docs = {'doc_a': None, 'doc_b': None, 'doc_c': None}
        remaining_pdfs = []
        for p in pdfs:
            doc_type = classify_document(os.path.basename(p))
            if doc_type and assigned_docs[doc_type] is None:
                assigned_docs[doc_type] = p
            else:
                remaining_pdfs.append(p)
        for key in ['doc_a', 'doc_b', 'doc_c']:
            if assigned_docs[key] is None and remaining_pdfs:
                assigned_docs[key] = remaining_pdfs.pop(0)

        extracted_docs = {}
        # Parallel Extraction Loop
        with concurrent.futures.ThreadPoolExecutor(max_workers=3) as executor:
            future_to_key = {}
            for key in ['doc_a', 'doc_b', 'doc_c']:
                pdf_file = assigned_docs.get(key)
                if pdf_file:
                    future_to_key[executor.submit(extract_shipping_details_llm, pdf_file)] = (key, pdf_file)
            
            for future in concurrent.futures.as_completed(future_to_key):
                key, pdf_file = future_to_key[future]
                row[f'{key}_Name'] = os.path.basename(pdf_file)
                try:
                    details = future.result()
                    extracted_docs[key] = {'details': details}
                    if details:
                        row[f'{key}_Cartons'] = details.get('cartons', {}).get('value')
                        row[f'{key}_Weight'] = details.get('gross_weight', {}).get('value')
                        row[f'{key}_Volume'] = details.get('cbm', {}).get('value')
                        
                        # Logic to rename BL file
                        if key == 'doc_a' and renamed_bls_dir and details.get('bl_number'):
                            try:
                                bl_num = "".join(c for c in details.get('bl_number') if c.isalnum() or c in ('-','_'))
                                if bl_num:
                                    ext = os.path.splitext(pdf_file)[1]
                                    new_name = f"{bl_num}{ext}"
                                    shutil.copy2(pdf_file, os.path.join(renamed_bls_dir, new_name))
                            except Exception as e:
                                print(f"Failed to rename BL: {e}")
                except Exception as e:
                    row['Error_Message'] += f"[{key} Err: {str(e)}] "
        
        # Compare
        comp_res = compare_three_documents(
            extracted_docs.get('doc_a', {}).get('details', {}),
            extracted_docs.get('doc_b', {}).get('details', {}),
            extracted_docs.get('doc_c', {}).get('details', {})
        )
        row['Status'] = 'MATCH' if comp_res.get('all_match') else 'MISMATCH'
        for comp in comp_res.get('comparisons', []):
                if comp['status'] != 'success':
                    row['Error_Message'] += f"{comp['field']} {comp['status']}; "
        
        return [row]

    except Exception as e:
        row['Status'] = 'Error'
        row['Error_Message'] = str(e)
        row['Duration_Seconds'] = round(time.time() - start_time, 2)
        return [row]
    finally:
        shutil.rmtree(temp_extract_dir)
        if 'Duration_Seconds' not in row:
             row['Duration_Seconds'] = round(time.time() - start_time, 2)


def process_combined_pdf(pdf_path, renamed_bls_dir=None):
    """
    Helper to process ONE combined PDF file.
    """
    start_time = time.time()
    row = {'Zip_Filename': os.path.basename(pdf_path), 'Status': '', 'Error_Message': ''}
    
    try:
        # Direct AI Logic
        extracted_docs = extract_combined_shipping_details_llm(pdf_path)
        
        # Populate row values for Excel
        for key in ['doc_a', 'doc_b', 'doc_c']:
             details = extracted_docs.get(key, {}).get('details', {})
             row[f'{key}_Name'] = "Combined"
             row[f'{key}_Cartons'] = details.get('cartons', {}).get('value')
             row[f'{key}_Weight'] = details.get('gross_weight', {}).get('value')
             row[f'{key}_Volume'] = details.get('cbm', {}).get('value')

        # Compare
        comp_res = compare_three_documents(
            extracted_docs.get('doc_a', {}).get('details', {}),
            extracted_docs.get('doc_b', {}).get('details', {}),
            extracted_docs.get('doc_c', {}).get('details', {})
        )
        row['Status'] = 'MATCH' if comp_res.get('all_match') else 'MISMATCH'
        
        for comp in comp_res.get('comparisons', []):
                if comp['status'] != 'success':
                    row['Error_Message'] += f"{comp['field']} {comp['status']}; "
        
        row['Duration_Seconds'] = round(time.time() - start_time, 2)
        return [row]

    except Exception as e:
        row['Status'] = 'Error'
        row['Error_Message'] = str(e)
        row['Duration_Seconds'] = round(time.time() - start_time, 2)
        return [row]


@app.route('/batch_process', methods=['POST'])
def batch_process():
    uploaded_files = request.files.getlist('zip_files')
    if not uploaded_files: return jsonify({'error': 'No files'}), 400

    job_id = str(uuid.uuid4())
    logger.info(f"Received batch request {job_id} with {len(uploaded_files)} files.")

    # Create persistent job dir
    job_dir = os.path.join(tempfile.gettempdir(), 'shipping_jobs', job_id)
    os.makedirs(job_dir, exist_ok=True)

    file_paths = []
    try:
        for f in uploaded_files:
            fname = f.filename.lower()
            if fname.endswith('.zip') or fname.endswith('.pdf'):
                # Save directly to disk, avoiding memory issues
                path = os.path.join(job_dir, f.filename) # insecure_filename technically, but trusted user
                f.save(path)
                file_paths.append(path)
    except Exception as e:
        logger.error(f"Error saving files for job {job_id}: {e}")
        return jsonify({'error': f'Failed to save files: {str(e)}'}), 500

    if not file_paths:
        return jsonify({'error': 'No valid ZIP or PDF files found'}), 400

    JOBS[job_id] = {'status': 'queued', 'progress': 0}
    
    # Start Thread
    thread = threading.Thread(target=process_batch_job, args=(job_id, file_paths, app))
    thread.daemon = True
    thread.start()
    
    return jsonify({'success': True, 'job_id': job_id})

@app.route('/batch_status/<job_id>', methods=['GET'])
def batch_status(job_id):
    job = JOBS.get(job_id)
    if not job: return jsonify({'error': 'Job not found'}), 404
    
    # Return a clean copy without binary data (bytes not proper for JSON)
    job_response = job.copy()
    job_response.pop('csv_data', None)
    job_response.pop('xlsx_data', None)
    job_response.pop('bl_zip_data', None)
    
    return jsonify(job_response)

@app.route('/batch_download/<job_id>', methods=['GET'])
def batch_download(job_id):
    job = JOBS.get(job_id)
    if not job or job['status'] != 'completed': return jsonify({'error': 'Not ready'}), 400
    
    if 'xlsx_data' in job:
        mem = io.BytesIO(job['xlsx_data'])
        mem.seek(0)
        return send_file(
            mem,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='batch_report.xlsx'
        )
    
    # Fallback/Legacy
    if 'csv_data' in job:
        mem = io.BytesIO()
        mem.write(job['csv_data'].encode('utf-8'))
        mem.seek(0)
        return send_file(
            mem,
            mimetype='text/csv',
            as_attachment=True,
            download_name='batch_report.csv'
        )
    
    return jsonify({'error': 'No report data'}), 404


@app.route('/batch_download_bls/<job_id>', methods=['GET'])
def batch_download_bls(job_id):
    job = JOBS.get(job_id)
    if not job or job['status'] != 'completed': return jsonify({'error': 'Not ready'}), 400
    
    if 'bl_zip_data' in job:
        mem = io.BytesIO(job['bl_zip_data'])
        mem.seek(0)
        return send_file(
            mem,
            mimetype='application/zip',
            as_attachment=True,
            download_name='renamed_bls.zip'
        )
    
    return jsonify({'error': 'No BL zip data found'}), 404


if __name__ == '__main__':
    app.run(debug=True, port=5000)
